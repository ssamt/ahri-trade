from openpyxl import load_workbook

from collections import defaultdict

book_f = open('books.txt', 'r', encoding='utf8')
content = book_f.read().splitlines()
books = []
price = []
for i in range(0, len(content), 2):
    books.append(content[i])
    price.append(int(content[i+1]))

def get_data(wb):
    names = [[] for i in range(len(books))]
    wb = wb['설문지 응답 시트1']
    wb = [[str(cell.value).strip() for cell in row] for row in wb]
    name_idx = dict()
    for i in range(1, len(wb)):
        name_idx[wb[i][1]] = i
    idx_name = {name_idx[k]:k for k in name_idx.keys()}
    for i in range(1, len(wb)):
        if i in idx_name:
            name = idx_name[i]
            all_books = ', '.join(wb[i][2:11])
            for j in range(len(books)):
                if books[j] in all_books:
                    names[j].append(name)
    return names

BUY, SELL = 0, 1
buy = load_workbook('구매 신청서(응답).xlsx')
buy = get_data(buy)
sell = load_workbook('판매 신청서(응답).xlsx')
sell = get_data(sell)
people = defaultdict(list)
for i in range(len(books)):
    print(books[i])
    print(f'구매: {buy[i]}')
    print(f'판매: {sell[i]}')
print('')
print('구매가 더 많음')
for i in range(len(books)):
    if len(buy[i]) > len(sell[i]):
        print(books[i])
print('판매가 더 많음')
for i in range(len(books)):
    if len(buy[i]) < len(sell[i]):
        print(books[i])
print('')
android = books.index('(영어독해와작문) Do Androids Dream of Electric Sheep?')
for i in range(len(books)):
    amount = min(len(buy[i]), len(sell[i]))
    for j in range(len(buy[i])):
        people[buy[i][j]].append([BUY, j<amount, i])
        if i == android and buy[i][j] == '21-006/국태영':
            people[buy[i][j]][-1][1] = True
        if i == android and buy[i][j] == '21-082/이진하':
            people[buy[i][j]][-1][1] = False
    for j in range(len(sell[i])):
        people[sell[i][j]].append([SELL, j<amount, i])
all_total = 0
total_transfer = 0
for p in people:
    print(p)
    for book in people[p]:
        if book[0] == BUY:
            if book[1]:
                print(f'[구매 O]: {books[book[2]]}')
            else:
                print(f'[구매 X]: {books[book[2]]}')
        else:
            if book[1]:
                print(f'[판매 O]: {books[book[2]]}')
            else:
                print(f'[판매 X]: {books[book[2]]}')
    total = 0
    for book in people[p]:
        if book[0] == BUY:
            if book[1]:
                total -= price[book[2]]//2
                print(f'{books[book[2]]}: -{price[book[2]]//2}원')
        else:
            if book[1]:
                total += price[book[2]]//2
                print(f'{books[book[2]]}: +{price[book[2]]//2}원')
                total_transfer += price[book[2]]//2
    print(f'총합: {total}원')
    all_total += total
print(all_total, total_transfer)
'''print('\n'.join(list(people.keys())))
own = defaultdict(int)
trade_f = open('trade.txt', 'r', encoding='utf8')
traded = trade_f.read().splitlines()
traded = [t.split() for t in traded]
print('')
for t in traded:
    if t[1] == 'O':
        person = t[0]
        for book in people[person]:
            if book[0] == BUY:
                if book[1]:
                    own[book[2]] -= 1
            else:
                if book[1]:
                    own[book[2]] += 1
    else:
        print(t[0])
print('')
for book_idx in own:
    print(books[book_idx], own[book_idx])
'''